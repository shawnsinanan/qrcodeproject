
# Hardcode the address? 
from segno import helpers 
# from PyPDF2 import PdfReader #For reading pdf-may be useful later
import openpyxl 
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import sys
# pip install xlwings

# Sub RunPythonScript()

# Dim objShell As Object
# Dim PythonExePath, PythonScriptPath As String

#     Set objShell = VBA.CreateObject("Wscript.Shell")
    
#     PythonExePath = """C:\Program Files\WindowsApps\PythonSoftwareFoundation.Python.3.9_3.9.3568.0_x64__qbz5n2kfra8p0"""
#     PythonScript Path = """C:\Users\tsuiwi\Downloads\winston.py"""
    
#     objShell.Run PythonExePath & PythonScriptPath

# End Sub














locate_python = sys.exec_prefix

print(locate_python)
#File should be read in through the program. Drag and drop in? Maybe user can select which excel file to input? Maybe a button?
data_file = 'Business Card Order #42.xlsx'  

wb = load_workbook(data_file, data_only=True)
# List all the sheets in the file.
# for sheetname in wb.sheetnames:
#     print(sheetname)

#Sets up a list of rows containing employee information. Works for one sheet only.
sheetnamelist=[]
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    all_rows = list(ws.rows)
    sheetname=wb.active #necessary later for setting row dimensions
    sheetnamelist.append(sheetname)

a=2
for row in all_rows[1:]:  #This changes the excel row and column dimensions so the QR Code can fit nicely.
     sheetnamelist[0].row_dimensions[a].height = 80
     a+=1
sheetnamelist[0].column_dimensions['K'].width = 18


i=0
tempforfile=""
for row in all_rows[1:]: #every row
    # print(rowheight)
    rowinfo=[]
    i+=1
    for cell in all_rows[i]:  #every cell in row

        
        rowinfo.append(cell.value)
    
    if (rowinfo[0]==None or rowinfo[0]=="" or rowinfo[0]==" "): #If the employee name is not entered or is a blank space, it will give an error. This fixes it.
        x=0
    else:
        qrcode = helpers.make_vcard(name = rowinfo[0][rowinfo[0].find(" "):] + ";"+rowinfo[0].split()[0], displayname=rowinfo[0], org="NYCDDC: "+rowinfo[3], url="nyc.gov/ddc",
                              email=rowinfo[7], street = "30-30 Thompson Ave.", city="Long Island City", region="NY", zipcode = "11101",
                               workphone=rowinfo[5], cellphone = rowinfo[6], title=rowinfo[2])  
        qrcodename=  rowinfo[0]+" QRCODE.png"      
        qrcode.save(qrcodename, scale=7)
        img = openpyxl.drawing.image.Image(qrcodename) #Adds qrcode to a new excel file. New excel file must not be in edit mode.
        img.anchor = "K"+str(i+1)
        img.height=110 
        img.width=110
        ws.add_image(img)
        wb.save("BBusiness Card Order #42.xlsx")
        print(rowinfo)



# print (type(rowinfo[5]))


# reader = PdfReader("Business Card Request.pdf")
# fields = reader.get_form_text_fields()
# testfields = reader.getFields()
# print (fields)
# print(testfields)

# try:  # In the case EmployeeName field is empty or has one word
#     firstname = fields["Employee Name"].split()[0]
# except:
#     firstname=""

# try: 
#     lastname = fields["Employee Name"].split()[1]
# except: 
#     lastname=""


# qrcode = helpers.make_vcard(name= lastname+ ";"+firstname, displayname=fields["Employee Name"], org="NYCDDC-"+fields["DivisionUnit"], url="nyc.gov/ddc",
#                              email=fields["Email Address"], street = "30-30 Thompson Ave.", city="Long Island City", region="NY", zipcode = "11101",
#                               workphone=fields["Office Phone"].strip(" /-\\"), phone = fields["Telephone Number"].strip(" /-\\"), cellphone = fields["Cellphone Number optional"].strip(" /-\\"), title=fields["Office Title"])
# qrcode.save("TESTQRCODE.png", scale=7)



# f = open("C:\\Users\\tsuiwi\\Desktop\\raft.txt", "r")
# print(f.readlines())

